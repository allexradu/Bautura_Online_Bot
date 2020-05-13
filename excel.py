import openpyxl
import pandas as pd
import platform
from openpyxl.styles import Alignment
import excel_data
import extra_functions

product_data = excel_data.ExcelData()

work_sheet_index = 2

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'


def save_product_to_excel():
    global work_sheet_index
    # loading the workbook
    wb = openpyxl.load_workbook(table_location)

    # wb.active returns a Worksheet object
    ws = wb.active
    # adding the data to the workbook
    main_category_key = extra_functions.value_key(product_data.main_category_cell_letter, work_sheet_index)
    ws[main_category_key] = product_data.main_category
    sub_category_key = extra_functions.value_key(product_data.sub_category_cell_letter, work_sheet_index)
    ws[sub_category_key] = product_data.sub_category
    product_title_key = extra_functions.value_key(product_data.product_title_cell_letter, work_sheet_index)
    ws[product_title_key] = product_data.product_title
    product_id_key = extra_functions.value_key(product_data.product_id_cell_letter, work_sheet_index)
    ws[product_id_key] = product_data.product_id
    price_text_key = extra_functions.value_key(product_data.price_text_cell_letter, work_sheet_index)
    ws[price_text_key] = product_data.price_text
    old_price_text_key = extra_functions.value_key(product_data.old_price_text_cell_letter, work_sheet_index)
    ws[old_price_text_key] = product_data.old_price_text
    description_key = extra_functions.value_key(product_data.description_cell_letter, work_sheet_index)
    ws[description_key] = product_data.description
    ws[description_key].alignment = Alignment(wrapText = True)
    alcohol_concentration_key = extra_functions.value_key(product_data.alcohol_concentration_cell_letter,
                                                          work_sheet_index)
    ws[alcohol_concentration_key] = product_data.alcohol_concentration
    brand_key = extra_functions.value_key(product_data.brand_cell_letter, work_sheet_index)
    ws[brand_key] = product_data.brand
    image_url_key = extra_functions.value_key(product_data.image_url_cell_letter, work_sheet_index)
    ws[image_url_key] = product_data.image_url

    # saving workbook
    wb.save(table_location)
